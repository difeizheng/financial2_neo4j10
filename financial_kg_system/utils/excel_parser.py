"""
Excel Parser Utility
Parses Excel files and converts them to node structures compatible with our graph system
"""

import openpyxl
from openpyxl.utils.cell import get_column_letter
from typing import List, Dict, Any, Optional
from ..services.excel_formula_engine import FormulaParser, ReferenceType


class ExcelParser:
    """Utility class to parse Excel files and extract cell information"""
    
    def __init__(self):
        self.formula_parser = FormulaParser()
    
    def parse_workbook_to_nodes(self, file_path: str) -> List[Dict[str, Any]]:
        """
        Parse an Excel workbook to a list of nodes compatible with our graph structure
        
        Args:
            file_path: Path to the Excel file to parse
            
        Returns:
            List of dictionaries containing cell information in the same format as node.json
        """
        workbook = openpyxl.load_workbook(file_path, data_only=False)  # Keep formulas
        all_nodes = []
        
        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            
            # Determine used range to efficiently parse cells
            rows = worksheet.max_row
            cols = worksheet.max_column
            
            for row in range(1, rows + 1):
                for col in range(1, cols + 1):
                    cell = worksheet.cell(row=row, column=col)
                    
                    # Skip empty cells unless they have formulas
                    if cell.value is None and cell.data_type != 'f':  # 'f' is for formula
                        continue
                    
                    # Generate cell ID
                    cell_id = f"{sheet_name}_{row}_{get_column_letter(col)}"
                    
                    # Parse formula if present
                    formula_raw = None
                    if cell.data_type == 'f' and cell.value is not None:  # It's a formula
                        formula_raw = str(cell.value)
                        # Normalize Excel formula formatting - strip leading =
                        if formula_raw.startswith('='):
                            formula_raw = formula_raw[1:]
                        
                        # Verify the formula is valid through our parser
                        try:
                            extracted_refs = self.formula_parser.extract_references(formula_raw)
                        except Exception:
                            # If formula is incorrectly formatted, still store it
                            pass
                    
                    # Determine if this is a header cell (by checking adjacent cells)
                    is_header = self._is_header_cell(worksheet, row, col)
                    
                    # Add to our nodes collection
                    cell_info = {
                        "id": cell_id,
                        "value": self._clean_cell_value(cell.value),
                        "formula_raw": formula_raw,
                        "sheet": sheet_name,
                        "is_head": is_header,
                        "row_category": self._get_row_category(worksheet, row, col),
                        "col_category": self._get_col_category(worksheet, row, col),
                        "row": row,
                        "col": get_column_letter(col)
                    }
                    
                    all_nodes.append(cell_info)
        
        return all_nodes
    
    def _clean_cell_value(self, value: Any) -> Optional[str]:
        """Clean up cell values for consistent storage"""
        if value is None:
            return None
        return str(value)
    
    def _is_header_cell(self, worksheet, row: int, col: int) -> bool:
        """Determine if a cell is a header based on position or value characteristics"""
        cell = worksheet.cell(row=row, column=col)
        
        # Check if adjacent cells in both directions contain more data
        # For example, if a cell has text and cells below or to the right also have content
        value = self._clean_cell_value(cell.value)
        if not value:
            return False
            
        # Look for text-like values (vs numbers) that are likely headers
        # Check if this looks like a label based on its content
        if not value.replace(" ", "").replace("_", "").replace("-", "").isalnum():
            # This has special characters, more likely a value than a label
            pass
        elif value.lower() in ['sum', 'total', 'average', 'count', 'subtotal']:
            return False  # These are results, not headers
        elif isinstance(cell.value, str) and cell.value.replace(" ", "").isalpha() and len(cell.value) <= 30:
            # Likely a header if it's text and not too long
            # Check if values below this cell are numeric
            below_numeric_count = 0
            for r in range(row + 1, min(row + 6, worksheet.max_row + 1)):  # Check next 5 rows
                below_cell = worksheet.cell(row=r, column=col)
                if isinstance(below_cell.value, (int, float)):
                    below_numeric_count += 1
            if below_numeric_count >= 2:
                return True  # If multiple values below are numeric, this is likely a header
        
        return False
    
    def _get_row_category(self, worksheet, row: int, col: int) -> Optional[str]:
        """Attempt to determine if this cell matches others in its row to identify categories"""
        first_col_value = self._clean_cell_value(worksheet.cell(row=row, column=1).value)
        
        # If the first column contains significant text, it might be a row identifier
        if first_col_value and len(first_col_value.strip()) > 0:
            return first_col_value
        return None
    
    def _get_col_category(self, worksheet, row: int, col: int) -> Optional[str]:
        """Attempt to determine if this cell matches others in its column to identify categories"""
        first_row_value = self._clean_cell_value(worksheet.cell(row=1, column=col).value)
        
        # If the first row contains significant text, it might be a category header 
        if first_row_value and len(first_row_value.strip()) > 0:
            return first_row_value
        return None


# Compatibility function name as referenced in the API layer
def parse_excel_to_nodes(file_path: str) -> List[Dict[str, Any]]:
    """
    Wrapper function to maintain compatibility with API layer references.
    
    Args:
        file_path: Path to the Excel file to parse
        
    Returns:
        List of dictionaries containing cell information in the same format as node.json
    """
    parser = ExcelParser()
    return parser.parse_workbook_to_nodes(file_path)