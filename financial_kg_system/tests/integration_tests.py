"""
Integration Tests for Financial Knowledge Graph System
Tests the complete system including Excel parsing, graph building, calculation, and API interaction
"""

import unittest
import tempfile
import os
from unittest.mock import patch, MagicMock
from neo4j import GraphDatabase
import openpyxl
from openpyxl import Workbook
from openpyxl.worksheet.table import Table, TableStyleInfo
import sys
from pathlib import Path

# Add project modules to path for imports
sys.path.insert(0, str(Path(__file__).parent.parent))

from services.excel_formula_engine import ExcelFormulaProcessor
from database.dependency_graph_builder import DependencyGraphBuilder
from services.smart_recalculation_engine import SmartRecalculationEngine
from utils.cache_manager import CacheManager, DualCacheDependencyManager
from utils.excel_parser import parse_excel_to_nodes


class TestFinancialKnowledgeGraphSystem(unittest.TestCase):
    """Integration tests for the complete financial knowledge graph system"""
    
    def setUp(self):
        """Set up test environment with mocked or actual components"""
        # Use a test Neo4j instance (assumes Neo4j is running locally for testing)
        self.neo4j_uri = os.getenv("TEST_NEO4J_URI", "bolt://localhost:7687")
        self.neo4j_user = os.getenv("TEST_NEO4J_USER", "neo4j")
        self.neo4j_password = os.getenv("TEST_NEO4J_PASSWORD", "password") 
        
        # Mock/Temp directory paths for temporary test files
        self.temp_dir = tempfile.mkdtemp()
        
        # Initialize components
        self.graph_builder = DependencyGraphBuilder(self.neo4j_uri, self.neo4j_user, self.neo4j_password)
        self.recalc_engine = SmartRecalculationEngine(self.neo4j_uri, self.neo4j_user, self.neo4j_password)
        self.cache_manager = CacheManager()
        self.dependency_manager = DualCacheDependencyManager(self.cache_manager)
        
        # Create sample nodes similar to node.json structure
        self.sample_nodes = [
            {
                "id": "参数输入表_4_D",
                "value": "1234.56",
                "formula_raw": None,
                "sheet": "参数输入表", 
                "is_head": False,
                "row_category": "工程计划",
                "col_category": "参数",
                "row": 4,
                "col": "D"
            },
            {
                "id": "参数输入表_5_D", 
                "value": "7890.12",
                "formula_raw": None,
                "sheet": "参数输入表",
                "is_head": False,
                "row_category": "成本估算",
                "col_category": "参数", 
                "row": 5,
                "col": "D"
            },
            {
                "id": "利润表_10_B", 
                "value": "1234.56+7890.12",
                "formula_raw": "参数输入表_4_D+参数输入表_5_D",  # Simplified reference format
                "sheet": "利润表",
                "is_head": False,
                "row_category": "营业收入",
                "col_category": "金额", 
                "row": 10,
                "col": "B"
            }
        ]
    
    def tearDown(self):
        """Clean up test environment"""
        # Clean up temp directory
        import shutil
        shutil.rmtree(self.temp_dir, ignore_errors=True)
        
        # Close connections
        self.graph_builder.close()
        self.recalc_engine.close()
    
    def test_01_excel_parsing_to_nodes(self):
        """Test the Excel parsing utility correctly converts Excel to node format"""
        # Create a minimal Excel file to parse
        excel_path = os.path.join(self.temp_dir, "test_input.xlsx")
        wb = Workbook()
        ws = wb.active
        ws.title = "参数输入表"
        
        # Add test data
        ws['A1'] = '类别'
        ws['B1'] = '参数'
        ws['C1'] = '数值'
        
        ws['A2'] = '工程计划'
        ws['B2'] = '建设期'
        ws['C2'] = 24  # months
        
        ws['A3'] = '成本估算'
        ws['B3'] = '总投资' 
        ws['C3'] = 1000000
        
        wb.save(excel_path)
        
        # Parse the Excel file
        parsed_nodes = parse_excel_to_nodes(excel_path)
        
        # Check that nodes were created correctly
        self.assertGreater(len(parsed_nodes), 0, "Should have parsed some nodes from Excel file")
        self.assertIn('参数输入表_2_A', [node['id'] for node in parsed_nodes])
        self.assertIn('参数输入表_2_B', [node['id'] for node in parsed_nodes])
        self.assertIn('参数输入表_2_C', [node['id'] for node in parsed_nodes])
        
        # Find and check the node with value '24'
        for node in parsed_nodes:
            if node['value'] == '24':
                self.assertEqual(node['id'], '参数输入表_3_C')
                self.assertEqual(node['sheet'], '参数输入表')
                break
        else:
            self.fail("Could not find node with value '24'")
    
    def test_02_graph_building(self):
        """Test importing nodes into Neo4j graph database"""
        # Import sample nodes
        model_id = self.graph_builder.import_workbook_structure(self.sample_nodes)
        
        self.assertIsNotNone(model_id)
        self.assertTrue(model_id.startswith("model_"))
        
        # Get specific cell details to verify it was imported properly
        cell_info = self.graph_builder.get_cell_details("参数输入表_4_D")
        self.assertIsNotNone(cell_info)
        self.assertEqual(cell_info.get('value_string'), "1234.56")
        self.assertEqual(cell_info.get('cell_type'), 'input')
    
    def test_03_formula_parsing(self):
        """Test Excel formula engine correctly parses and identifies dependencies"""
        # Use the ExcelFormulaProcessor to identify dependencies
        processor = ExcelFormulaProcessor()
        
        # Example formula with dependencies
        formula_with_deps = "参数输入表_4_D+参数输入表_5_D"
        dependencies = processor.get_dependencies(formula_with_deps, "利润表")
        
        # Check that dependencies were found correctly
        dependency_ids = [dep[0] for dep in dependencies]
        self.assertIn("参数输入表_4_D", dependency_ids)
        self.assertIn("参数输入表_5_D", dependency_ids)
    
    def test_04_dependency_detection(self):
        """Test dependency graph builder correctly identifies cell relations"""
        # First ensure sample data is in DB 
        model_id = self.graph_builder.import_workbook_structure(self.sample_nodes)
        
        # Find dependencies of '利润表_10_B' - should include our two parameter cells
        dependents_of_param1 = self.graph_builder.find_dependent_cells(["参数输入表_4_D"])
        self.assertIn("利润表_10_B", dependents_of_param1)
    
    def test_05_smart_recalculation(self):
        """Test the recaclculation engine updates values correctly"""
        # Import a slightly more complex example with real dependencies
        complex_nodes = [
            {
                "id": "参数表_2_A",
                "value": "100",
                "formula_raw": None,
                "sheet": "参数表",
                "is_head": False,
                "row_category": None,
                "col_category": None,
                "row": 2,
                "col": "A"
            },
            {
                "id": "参数表_2_B", 
                "value": "200",
                "formula_raw": None,
                "sheet": "参数表",
                "is_head": False,
                "row_category": None,
                "col_category": None,
                "row": 2,
                "col": "B"
            },
            {
                "id": "计算表_3_A",
                "value": "=A2+B2",
                "formula_raw": "参数表_2_A+参数表_2_B",
                "sheet": "计算表",  # Adjust for internal reference system
                "is_head": False, 
                "row_category": None,
                "col_category": None,
                "row": 3,
                "col": "A"
            }
        ]
        
        # Import into the graph
        model_id = self.graph_builder.import_workbook_structure(complex_nodes)
        
        # Change parameter and recalculate
        cell_change = {"参数表_2_A": 150}  # Change from 100 to 150
        result = self.recalc_engine.calculate_from_changes(cell_change)
        
        # The result should include the calculation table cell since it depends on param
        self.assertIn("计算表_3_A", result.affected_cells)

    def test_06_cache_functionality(self):
        """Test the cache management system works correctly"""
        # Test graph caching
        import networkx as nx
        
        # Create a simple test graph
        test_graph = nx.DiGraph()
        test_graph.add_node("A")
        test_graph.add_node("B")
        test_graph.add_edge("A", "B")  # A depends on B
        
        # Cache it
        success = self.cache_manager.cache_graph("test_graph_1", test_graph)
        self.assertTrue(success, "Should successfully cache a graph")
        
        # Retrieve it
        retrieved_graph = self.cache_manager.get_cached_graph("test_graph_1")
        self.assertIsNotNone(retrieved_graph, "Should retrieve cached graph")
        self.assertEqual(len(retrieved_graph.nodes()), 2, "Should have same number of nodes")
        self.assertEqual(len(retrieved_graph.edges()), 1, "Should have same number of edges")
        
        # Test value caching
        values_to_cache = {"cell_X": 123.45, "cell_Y": 678.90}
        success = self.cache_manager.cache_cell_values(values_to_cache, ttl=60)
        self.assertTrue(success, "Should successfully cache cell values")
        
        retrieved_values = self.cache_manager.get_cached_cell_values(["cell_X", "cell_Y"])
        self.assertEqual(retrieved_values.get("cell_X"), 123.45)
        self.assertEqual(retrieved_values.get("cell_Y"), 678.90)
    
    def test_07_impact_analysis(self):
        """Test the impact analysis function works correctly"""
        # Set up dependency structure
        model_id = self.graph_builder.import_workbook_structure(self.sample_nodes)
        
        # Analyze impact of changing one cell
        analysis = self.recalc_engine.get_impact_analysis("参数输入表_4_D")
        
        # Should show that changing param cell affects profit cell
        self.assertIn("利润表_10_B", analysis.get("all_affected_cells", []))
        self.assertGreaterEqual(analysis.get("total_dependencies", 0), 1)
    
    def test_08_full_integration_workflow(self):
        """Test a complete end-to-end workflow: upload Excel, make changes, recalculate"""
        # 1. Create a test Excel file with formulas
        excel_path = os.path.join(self.temp_dir, "integration_test.xlsx")
        wb = Workbook()
        
        # Create input sheet
        input_ws = wb.create_sheet("输入参数")
        input_ws['A1'] = '项目'
        input_ws['B1'] = '金额'
        input_ws['A2'] = '收入'
        input_ws['B2'] = 1000  # Base value will be modified later
        input_ws['A3'] = '成本'
        input_ws['B3'] = 400
        
        # Create calculation sheet
        calc_ws = wb.create_sheet("计算结果")
        calc_ws['A1'] = '利润'
        calc_ws['B1'] = '=输入参数.B2-输入参数.B3'  # Simplified formula format
        
        wb.save(excel_path)
        
        # 2. Parse the Excel to nodes
        nodes = parse_excel_to_nodes(excel_path)
        self.assertGreater(len(nodes), 0, "Should parse test Excel file")
        
        # 3. Import into graph database
        model_id = self.graph_builder.import_workbook_structure(nodes)
        self.assertIsNotNone(model_id)
        
        # 4. Perform a recalculation after changing a parameter
        change_request = {"输入参数_2_B": 1200}  # Change income from 1000 to 1200
        calc_result = self.recalc_engine.calculate_from_changes(change_request)
        
        # Verify something was changed (our dependent calculation)
        self.assertGreater(len(calc_result.affected_cells), 0)
        # Note: Actual value checking would require deeper Excel formula execution
        # which would be done with xlwings in a real system


def run_integration_tests():
    """Run all integration tests"""
    suite = unittest.TestLoader().loadTestsFromTestCase(TestFinancialKnowledgeGraphSystem)
    runner = unittest.TextTestRunner(verbosity=2)
    result = runner.run(suite)
    
    return result.wasSuccessful()


if __name__ == '__main__':
    # This test suite can be run independently
    success = run_integration_tests()
    exit(0 if success else 1)